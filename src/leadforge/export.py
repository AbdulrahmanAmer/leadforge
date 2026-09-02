"""Export (U6.1-6.3, v0.3 U9.D): styled XLSX (Leads/Summary/About) + CSV mirror + report.json (docs/03 §5,
docs/09 "D — Scoring and export truth").

openpyxl only (pure-python). CSV is utf-8-sig so Excel on Windows opens it clean. Never prints data rows.
"""

from __future__ import annotations

import csv
import json
import re
import sqlite3
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from leadforge import compliance, db
from leadforge.config import Config
from leadforge.enrich.validate import rank_email_contacts
from leadforge.models import ICP
from leadforge.score import fill_email_affinity, phone_is_validated
from leadforge.util import natural_name, now_iso

# export_run's `cfg` parameter is optional (existing call sites in cli.py/pipeline.py pass none) — when
# omitted, the freemail_policy/require_corporate that feed Lawful Basis/Next Action fall back to this
# built-in default (same defaults score.py's contactability/status use when it isn't handed a Config
# either — see score.py's _DEFAULT_POLICY). A campaign with a non-default leadforge.yaml should pass its
# real `cfg` through to see that reflected in the sheet.
_DEFAULT_CFG = Config()

COLUMNS = [
    "Score", "Tier", "Business", "Category", "DM Name", "DM Title", "DM Conf", "Phone",
    "Email", "Email Tier", "Email (Inferred)", "Website", "Address", "City", "Region", "Postal", "Country",
    "Rating", "Reviews", "Likely Need (Hook)", "Why This Score", "Maps", "Source", "Verified On", "Stale?",
    "Opening Hours", "Company No", "Incorporated", "Company Status", "SIC Codes", "Call Readiness",
]
# account_fit profile (WE SCORE spec) appends the account-intel columns
ACCOUNT_COLUMNS = COLUMNS + [
    "Employees", "Employee Range", "Revenue", "Departments", "Microsoft 365", "CRM", "ERP",
    "Other Systems", "Trigger", "Trigger Strength", "LinkedIn", "Contactability", "Data Confidence",
    "Status",
]
# default profile (v0.3) appends the truth/compliance columns: fit vs contactability split, the
# phone-first Next Action, and the compliance facts a human decides outreach from.
DEFAULT_EXTRA_COLUMNS = [
    "Fit", "Contactability", "Status", "Next Action", "Entity Type", "Lawful Basis (Email)",
    "Registry Name", "Registry Match", "Chain", "Site Status", "Email Confidence", "All Hooks",
]
DEFAULT_COLUMNS = COLUMNS + DEFAULT_EXTRA_COLUMNS
_TIER_FILL = {
    "A": PatternFill("solid", fgColor="C6EFCE"),
    "B": PatternFill("solid", fgColor="FFEB9C"),
    "C": PatternFill("solid", fgColor="FCE4D6"),
    "D": PatternFill("solid", fgColor="E7E6E6"),
    "DQ": PatternFill("solid", fgColor="D9D9D9"),
}
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REGION_REMINDER = {
    "us": "US/CAN-SPAM: identify yourself, include a physical postal address + working opt-out; honor within 10 business days.",
    "uk": "UK/PECR: corporate subscribers may be emailed B2B; sole traders/individuals need consent. Always give identity + opt-out.",
    "eu": "EU/GDPR: legitimate-interest basis; keep the campaign LIA note; honor objections immediately; store source per contact.",
}


# chars openpyxl refuses (IllegalCharacterError) — scraped text can contain them
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _safe_cell(v):
    """Excel hardening for scraped strings: strip control chars openpyxl crashes on, and neutralize
    leading formula triggers so a hostile business name can't execute in the operator's Excel.
    '+' and '-' are deliberately left alone — spaced phone display and '-' placeholders are
    verified text-safe, and neither starts a formula when a cell contains further spaces."""
    if not isinstance(v, str):
        return v
    v = _ILLEGAL_XLSX_RE.sub("", v)
    if v[:1] in ("=", "@"):
        v = "'" + v
    return v


def _display_phone(e164: str | None) -> str:
    """"+44 1483 456363", not "+441483456363": Excel coerces a leading-+ digit run to a number
    (4.41483E+11); the spaced international format survives as text in both XLSX and CSV."""
    if not e164:
        return ""
    try:
        import phonenumbers
        return phonenumbers.format_number(phonenumbers.parse(e164, None),
                                          phonenumbers.PhoneNumberFormat.INTERNATIONAL)
    except Exception:  # noqa: BLE001 — display formatting must never lose the number
        return e164


def _site_status(enrich: dict) -> tuple[str, bool]:
    """-> ('live' | 'not crawlable (robots)' | 'dead (<status/error>)' | 'redirects to <host>' |
    'not crawled', site_dead). Reads signals C1/A set when present (final_host, offsite_redirect,
    http_status); degrades to 'live' on a bare crawl with none of those (nothing observed to say
    otherwise). robots-disallowed means the crawler was refused, not that the site is down — it is
    reported honestly and does NOT set site_dead (which would otherwise withdraw email eligibility
    from a live site for a reason that was never actually observed)."""
    if not enrich.get("crawled_at"):
        return "not crawled", False
    signals = enrich.get("signals") or {}
    err = enrich.get("error")
    status = signals.get("http_status")
    if err == "robots-disallowed":
        return "not crawlable (robots)", False
    dead = bool(err) or (isinstance(status, int) and status >= 400)
    if dead:
        return f"dead ({status if status is not None else err})", True
    if signals.get("offsite_redirect") and signals.get("final_host"):
        return f"redirects to {signals['final_host']}", False
    return "live", False


def _email_confidence(affinity: str, tier: str, linkage_checked: bool = True) -> str:
    """Human-readable affinity+tier, per docs/09 §D. 'none' when there is no sendable candidate at all.
    linkage_checked=False means the affinity came from the coarse pre-v0.3 fallback (ANY freemail on
    the domain), not the real name/business-token linkage check — the column must not claim a check
    that never ran."""
    if affinity == "own_domain":
        return f"own domain, {tier} mailbox" if tier else "own domain"
    if affinity == "freemail_linked":
        return ("personal freemail, linked to owner name" if linkage_checked
                else "personal freemail (linkage not checked — pre-v0.3 row)")
    if affinity == "freemail_unlinked":
        return "personal freemail, UNLINKED — do not mail"
    if affinity == "foreign":
        return "foreign domain — not the business's own"
    return "none"


def _row_for(conn: sqlite3.Connection, s, staleness_days: int = 90, *,
            icp: ICP | None = None, cfg: Config | None = None, chain: dict | None = None) -> dict:
    cfg = cfg or _DEFAULT_CFG
    profile = icp.scoring.profile if icp else "default"
    people = db.people_for(conn, s["business_id"])
    dm = next((p for p in people if p["is_dm"] == 1), None)
    contacts = db.contacts_for(conn, s["business_id"])
    # Affinity MUST be backfilled BEFORE ranking (docs/09 §D fix): every pre-v0.3 row stores
    # affinity '' (100% of the live campaign DB), and this SAME filled list is what
    # compliance.email_eligibility below ranks too, so Email / Email Confidence / Lawful Basis /
    # contactability can never describe two different addresses for one business.
    contacts_filled = fill_email_affinity(contacts, s["domain"])
    # Best email: affinity-first (own-domain always outranks a freemail box, even a 'valid' one — the
    # v0.2 sheet exported a font designer's gmail above a real info@ three times), tier second.
    # 'inferred' addresses are NEVER candidates for the Email column — they get their own,
    # so a mail-merge over 'Email' can never pick up a guess.
    ranked_emails = rank_email_contacts(contacts_filled)
    sendable_ranked = [c for c in ranked_emails if c["tier"] not in ("invalid", "inferred")]
    best_row = sendable_ranked[0] if sendable_ranked else None
    best_email = (best_row["value"], best_row["tier"] or "") if best_row else ("", "")
    best_affinity = (best_row["affinity"] or "") if best_row else ""
    best_affinity_backfilled = bool(best_row and best_row.get("_affinity_backfilled"))
    inferred = next((c for c in contacts_filled if c["kind"] == "email" and c["tier"] == "inferred"), None)
    factors = json.loads(s["factors_json"])
    # 'Why This Score' explains Score/Fit — the meta factors (contactability, status) are a separate
    # axis by design (docs/09 §D split) and must never re-blend into it: contactability alone can carry
    # up to 98 points, which would silently outrank every fit factor (max 25) if left in this sort.
    top_why = "; ".join(f["why"] for f in sorted((f for f in factors if f.get("group") != "meta"),
                                                  key=lambda f: -f["points"])[:3])
    hooks = json.loads(s["need_hooks_json"])
    verified = ""
    ev = db.evidence_for(conn, s["business_id"])
    if ev:
        verified = max((e["observed_at"] for e in ev), default="")
    row = {
        "Score": round(s["total"]), "Tier": s["tier"], "Business": s["name"], "Category": s["category"] or "",
        "DM Name": natural_name(dm["name"]) if dm else "", "DM Title": dm["title"] if dm else "",
        "DM Conf": round(dm["dm_confidence"], 2) if dm else "",
        "Phone": _display_phone(s["phone_e164"]) or s["phone_raw"]
                 or _display_phone(next((c["value"] for c in contacts if c["kind"] == "phone"), None)),
        "Email": best_email[0], "Email Tier": best_email[1],
        "Website": s["website"] or "", "Address": s["address_street"] or s["address_full"] or "",
        "City": s["address_city"] or "", "Region": s["address_region"] or "", "Postal": s["address_postal"] or "",
        "Country": s["address_country"] or "", "Rating": s["rating"] if s["rating"] is not None else "",
        "Reviews": s["review_count"] if s["review_count"] is not None else "",
        "Likely Need (Hook)": hooks[0] if hooks else "", "Why This Score": top_why,
        "Maps": s["maps_url"] or "", "Source": s["source"] or "", "Verified On": verified,
        "Stale?": _stale_flag(verified, staleness_days),
    }
    enrich_all = json.loads(s["enrich_json"]) if s["enrich_json"] else {}
    regp = enrich_all.get("registry_profile") or {}
    crawled = bool(enrich_all.get("crawled_at"))
    # Honesty markers for the Summary/report stats, captured BEFORE the placeholder text below
    # fills the cells — "not identified - ask for owner/manager" must never count as a DM.
    # Underscore keys are stripped by the writers (they only emit `columns`).
    row["_has_email"] = bool(row["Email"])
    row["_has_dm"] = bool(row["DM Name"])
    row["_has_inferred"] = inferred is not None
    row["_has_site"] = bool(row["Website"])
    if inferred is not None:
        imeta = json.loads(inferred["meta_json"]) if inferred["meta_json"] else {}
        conf = imeta.get("confidence")
        # the cell carries its own caveat: this is a likely address, not an observed one
        row["Email (Inferred)"] = (f"{inferred['value']} (likely, "
                                   f"{int(float(conf) * 100)}% — {imeta.get('basis', 'pattern')})"
                                   if conf else inferred["value"])
    else:
        row["Email (Inferred)"] = "not inferred"
    # No unresolved cells: a blank is replaced by WHY it is blank, so callers know what they hold.
    if not row["Email"]:
        row["Email"] = "none published" if crawled else ("no website to crawl" if not row["Website"] else "site not crawled")
        row["Email Tier"] = "-"
    if not row["Website"]:
        row["Website"] = "NONE - no web presence (pitch opportunity)"
    if not row["DM Name"]:
        row["DM Name"] = "not identified - ask for owner/manager"
    row["Opening Hours"] = _format_hours(s["hours_json"])
    row["Company No"] = regp.get("company_number") or ("not matched in registry" if enrich_all.get("registry_checked") else "not looked up")
    row["Incorporated"] = (regp.get("incorporated") or "")[:4] or "-"
    row["Company Status"] = regp.get("company_status") or "-"
    row["SIC Codes"] = ", ".join(regp.get("sic_codes") or []) or "-"
    # READY requires a VALIDATED number (parsed e164 or a site-extracted phone) — raw Maps
    # strings display in the Phone cell but are not call-ready evidence
    validated_phone = bool(s["phone_e164"]) or any(c["kind"] == "phone" for c in contacts)
    has_dm = row["_has_dm"]
    row["Call Readiness"] = ("READY - named contact" if validated_phone and has_dm
                             else "READY - ask switchboard" if validated_phone
                             else "UNVERIFIED PHONE - confirm number" if row["Phone"]
                             else "NO PHONE - research first")
    meta = {f["factor"]: f for f in factors if f.get("group") == "meta"}
    if profile == "account_fit":  # WE SCORE profile -> append the account-intel columns
        enrich = json.loads(s["enrich_json"]) if s["enrich_json"] else {}
        prof = enrich.get("profile") or {}
        tech = prof.get("tech") or {}
        trig = (prof.get("triggers") or [{}])[0]
        socials = enrich.get("socials") or {}

        def tri(key):
            f = tech.get(key) or {}
            v = f.get("value")
            return (f.get("name") or "yes") if v == "yes" else ("no" if v == "no" else "unknown")

        row.update({
            "Employees": (prof.get("employee_count") or {}).get("value") or "",
            "Employee Range": prof.get("employee_range", "unknown"),
            "Revenue": (prof.get("revenue") or {}).get("value") or "unknown",
            "Departments": ", ".join(prof.get("departments") or []),
            "Microsoft 365": tri("microsoft_365"), "CRM": tri("crm"), "ERP": tri("erp"),
            "Other Systems": ", ".join(tech.get("other") or []),
            "Trigger": trig.get("text", ""), "Trigger Strength": trig.get("strength", ""),
            "LinkedIn": socials.get("linkedin", ""),
            "Contactability": meta["contactability"]["points"] if "contactability" in meta else "",
            "Data Confidence": meta["data_confidence"]["points"] if "data_confidence" in meta else "",
            "Status": meta["status"]["why"],
        })
        return row

    # v0.3 default-profile columns: the fit/contactability split, phone-first Next Action, and the
    # compliance facts a human decides outreach from. Computed fresh here (not read off the stored
    # Score) because entity/eligibility/outreach-state can change between scoring and export.
    entity = compliance.entity_type(s, people)
    site_status, site_dead = _site_status(enrich_all)
    is_suppressed = db.is_suppressed(conn, s["domain"], s["place_id"], best_email[0] or None)
    eligibility = compliance.email_eligibility(
        s, contacts_filled, entity, icp.compliance.region_profile if icp else "us",
        freemail_policy=cfg.validation.freemail_policy, require_corporate=cfg.outreach.require_corporate,
        suppressed=is_suppressed, site_dead=site_dead,
    )
    phone_ok = phone_is_validated(s["phone_e164"])
    next_action = compliance.next_action(
        phone_validated=phone_ok, has_dm=row["_has_dm"], eligibility=eligibility, tier=s["tier"],
        outreach_state=db.outreach_state_for(conn, s["business_id"]),
    )
    chain_key = (chain or {}).get(s["business_id"])
    sim = regp.get("match_similarity")
    row["_email_affinity"] = best_affinity
    row["_eligible"] = bool(eligibility.get("eligible"))
    row["_phone_validated"] = phone_ok
    row.update({
        "Fit": row["Score"],
        "Contactability": meta["contactability"]["points"] if "contactability" in meta else "not computed",
        "Status": meta["status"]["why"] if "status" in meta else "not computed",
        "Next Action": next_action,
        "Entity Type": entity,
        "Lawful Basis (Email)": eligibility.get("basis", compliance.BASIS_NONE),
        "Registry Name": regp.get("legal_name")
                         or ("not matched in registry" if enrich_all.get("registry_checked") else "not looked up"),
        "Registry Match": f"{float(sim):.2f}" if sim is not None else "n/a",
        "Chain": chain_key or "-",
        "Site Status": site_status,
        "Email Confidence": _email_confidence(best_affinity, best_email[1],
                                              linkage_checked=not best_affinity_backfilled)
                           if best_email[0] else "none",
        "All Hooks": "; ".join(hooks) if hooks else "no hooks triggered",
    })
    return row


def export_run(conn: sqlite3.Connection, icp: ICP, run_id: str, out_dir: Path, formats: list[str],
               staleness_days: int = 90, cfg: Config | None = None) -> list[str]:
    cfg = cfg or _DEFAULT_CFG
    profile = icp.scoring.profile
    rows_raw = db.scores_for_run(conn, run_id)
    chain = db.chain_map(conn) if profile != "account_fit" else {}
    rows = [_row_for(conn, s, staleness_days, icp=icp, cfg=cfg, chain=chain) for s in rows_raw]
    rows.sort(key=lambda r: (-r["Score"], r["Tier"]))
    columns = ACCOUNT_COLUMNS if profile == "account_fit" else DEFAULT_COLUMNS
    run_dir = out_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    artifacts: list[str] = []

    if "xlsx" in formats:
        artifacts.append(str(_write_xlsx(run_dir / f"{icp.campaign}.xlsx", rows, icp, run_id, columns)))
    if "csv" in formats:
        artifacts.append(str(_write_csv(run_dir / f"{icp.campaign}.csv", rows, columns)))
    artifacts.append(str(_write_report(run_dir / "report.json", rows, icp, run_id)))
    return artifacts


def _write_csv(path: Path, rows: list[dict], columns: list[str] = COLUMNS) -> Path:
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.DictWriter(fh, fieldnames=columns, restval="-")  # same blank rule as the XLSX
        w.writeheader()
        for r in rows:
            w.writerow({k: _safe_cell(v if v != "" else "-") for k, v in r.items() if k in columns})
    return path


def _write_xlsx(path: Path, rows: list[dict], icp: ICP, run_id: str,
                columns: list[str] = COLUMNS) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(columns)
    for ci in range(1, len(columns) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([_safe_cell(r.get(c, '') if r.get(c, '') != '' else '-') for c in columns])
    # styling: tier fill, hyperlinks, widths (deliberately no zebra — tier colors carry the rows)
    tier_col = columns.index("Tier") + 1
    web_col = columns.index("Website") + 1
    maps_col = columns.index("Maps") + 1
    phone_col = columns.index("Phone") + 1
    for ri in range(2, len(rows) + 2):
        ws.cell(row=ri, column=phone_col).number_format = "@"  # text — never scientific notation
        tier = ws.cell(row=ri, column=tier_col).value
        fill = _TIER_FILL.get(tier)
        if fill:
            ws.cell(row=ri, column=tier_col).fill = fill
        for col in (web_col, maps_col):
            cell = ws.cell(row=ri, column=col)
            # only real URLs get link styling — '-' and 'NONE - no web presence' placeholders
            # used to render as blue clickable links that opened a relative path named '-'
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    _autosize(ws)

    _summary_sheet(wb, rows, icp, run_id)
    _about_sheet(wb, icp)
    wb.save(path)
    return path


def _summary_sheet(wb: Workbook, rows: list[dict], icp: ICP, run_id: str) -> None:
    ws = wb.create_sheet("Summary")
    tiers = {t: sum(1 for r in rows if r["Tier"] == t) for t in ("A", "B", "C", "D", "DQ")}
    hook_counts: dict[str, int] = {}
    for r in rows:
        if r["Likely Need (Hook)"]:
            hook_counts[r["Likely Need (Hook)"]] = hook_counts.get(r["Likely Need (Hook)"], 0) + 1
    with_dm, with_email = _real_counts(rows)
    lines = [
        ("LeadForge campaign", icp.campaign),
        ("Offer", icp.offer.what),
        ("Categories", ", ".join(icp.target.categories)),
        ("Geography", ", ".join(icp.target.geography.areas) or "bbox"),
        ("Run", run_id),
        ("Generated", now_iso()),
        ("", ""),
        ("Total leads", len(rows)),
        ("Tier A / B / C / D / DQ",
         f"{tiers['A']} / {tiers['B']} / {tiers['C']} / {tiers['D']} / {tiers['DQ']}"),
        ("With decision maker", f"{with_dm} ({_pct(with_dm, len(rows))})"),
        ("With email (published)", f"{with_email} ({_pct(with_email, len(rows))})"),
        ("With inferred email (guess)", str(sum(1 for r in rows if r.get("_has_inferred")))),
        ("", ""),
        ("Top need hooks", ""),
    ]
    for hook, n in sorted(hook_counts.items(), key=lambda x: -x[1])[:8]:
        lines.append((f"  {n}×", hook))
    # v0.3 default-profile funnel + Next Action breakdown (rows carry these keys only when the
    # default-profile columns were computed — account_fit rows skip this block cleanly)
    if any("Next Action" in r for r in rows):
        with_site = sum(1 for r in rows if r.get("_has_site"))
        with_own_domain = sum(1 for r in rows if r.get("_email_affinity") == "own_domain")
        with_eligible = sum(1 for r in rows if r.get("_eligible"))
        call_ready = sum(1 for r in rows if r.get("_phone_validated"))
        lines += [
            ("", ""),
            ("Funnel: sites -> any email -> own-domain email -> eligible to email -> call-ready", ""),
            ("  With website", f"{with_site} ({_pct(with_site, len(rows))})"),
            ("  With any email", f"{with_email} ({_pct(with_email, len(rows))})"),
            ("  With own-domain email", f"{with_own_domain} ({_pct(with_own_domain, len(rows))})"),
            ("  Eligible to email", f"{with_eligible} ({_pct(with_eligible, len(rows))})"),
            ("  Call-ready (validated phone)", f"{call_ready} ({_pct(call_ready, len(rows))})"),
            ("", ""),
            ("Next Action breakdown", ""),
        ]
        na_counts: dict[str, int] = {}
        for r in rows:
            na = r.get("Next Action", "")
            if na:
                na_counts[na] = na_counts.get(na, 0) + 1
        for na, n in sorted(na_counts.items(), key=lambda x: -x[1]):
            lines.append((f"  {n}×", na))
    lines += [("", ""), ("Compliance reminder", _REGION_REMINDER.get(icp.compliance.region_profile, ""))]
    for k, v in lines:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    ws["A1"].font = Font(bold=True, size=14)


def _about_sheet(wb: Workbook, icp: ICP) -> None:
    ws = wb.create_sheet("About")
    ws.append(["LeadForge — how to read this sheet"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    # the legend must match the rubric that actually graded this workbook (score.py _grade
    # for account_fit; tiers block of scoring.default.yaml otherwise)
    account_fit = getattr(icp.scoring, "profile", "") == "account_fit"
    if account_fit:
        ws.append(["Tier", "A ≥ 80 · B 65–79 · C 50–64 · D < 50 · DQ = hard-disqualified by your ICP rules"])
    else:
        ws.append(["Tier", "A ≥ 80 · B 60–79 · C < 60 · DQ = hard-disqualified by your ICP rules "
                            "(tier is FIT only — see Contactability/Status for reachability)"])
    ws.append(["Score", "0–100, sum of weighted factors; see 'Why This Score' per row for the top drivers"])
    ws.append(["Email Tier", "valid > role > risky > catch_all > unknown (invalid emails are dropped)"])
    ws.append(["Email (Inferred)", "NOT a published address: the address this domain's own naming "
                                   "convention implies for the named contact (derived from a real "
                                   "email already found there). Confirm before use; never counted "
                                   "in the 'with email' figure."])
    ws.append(["DM Conf", "0–1 confidence the agent assigned when labeling the decision maker from site snippets"])
    ws.append(["Likely Need (Hook)", "auto-suggested outreach angle from detected need signals + your offer"])
    ws.append(["Verified On", "timestamp of the newest evidence for the row; older than your staleness window = re-verify"])
    ws.append(["Source", "which engine discovered the business (gosom = Google Maps)"])
    if not account_fit:
        ws.append([])
        ws.append(["v0.3 columns", "fit/contactability are graded separately on purpose — a hot-but-"
                                    "uncontactable lead and a lukewarm-but-call-ready one are never blended."])
        ws.append(["Fit", "0–100, industry/need/size/geography/business-model/data-confidence — same "
                          "number as Score, drives Tier"])
        ws.append(["Contactability", "0–100, separate: DM + best email + validated phone + registry "
                                     "corroboration + mobile — never affects Tier"])
        ws.append(["Status", "READY (tier A/B and contactability ≥ 50) · CALL_ONLY (validated phone, "
                             "no eligible email) · RESEARCH (neither yet) · DQ (hard-disqualified)"])
        ws.append(["Next Action", "phone-first: CALL a named contact, or the switchboard, before EMAIL; "
                                  "shows the live outreach state once a lead is enrolled in a campaign"])
        ws.append(["Entity Type", "what the public company registry says: corporate_active/_inactive/"
                                  "_unknown, unmatched (likely sole trader), or unchecked (no lookup run)"])
        ws.append(["Lawful Basis (Email)", "the basis an unsolicited email would rest on under your "
                                           "region's rules and freemail policy — 'none' = do not email"])
        ws.append(["Registry Name / Match", "the matched company's legal name and the name-similarity "
                                            "score (0–1) that accepted the match; 'n/a' = no match"])
        ws.append(["Chain", "shares a domain or phone with another row in this database (same operator, "
                            "multiple locations) — '-' = independent as far as this data shows"])
        ws.append(["Site Status", "live / dead (<code or error>) / redirects to <host> / not crawled"])
        ws.append(["Email Confidence", "plain-language affinity + tier for the exported Email address"])
        ws.append(["All Hooks", "every need signal that fired, not just the top one shown in the hook column"])
    ws.append([])
    ws.append(["Note", "Contact data is public-source and probabilistic. Confirm before high-stakes outreach; honor opt-outs via `leadforge suppress add`."])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def _real_counts(rows: list[dict]) -> tuple[int, int]:
    """(with_dm, with_email) counting only real data — placeholder cells don't inflate coverage.
    Inferred addresses are deliberately excluded: they are guesses, counted separately."""
    return (sum(1 for r in rows if r.get("_has_dm")), sum(1 for r in rows if r.get("_has_email")))


def _write_report(path: Path, rows: list[dict], icp: ICP, run_id: str) -> Path:
    with_dm, with_email = _real_counts(rows)
    report = {
        "run": run_id, "campaign": icp.campaign, "generated": now_iso(), "total": len(rows),
        "tiers": {t: sum(1 for r in rows if r["Tier"] == t) for t in ("A", "B", "C", "D", "DQ")},
        "with_dm": with_dm,
        "with_email": with_email,
        "with_inferred_email": sum(1 for r in rows if r.get("_has_inferred")),
    }
    path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return path


def _autosize(ws, cap: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[letter].width = min(cap, max(10, width + 2))


def _pct(n: int, total: int) -> str:
    return f"{(100 * n / total):.0f}%" if total else "0%"


def summarize_for_digest(conn: sqlite3.Connection, run_id: str) -> dict:
    rows = db.scores_for_run(conn, run_id)
    tiers = {t: 0 for t in ("A", "B", "C", "D", "DQ")}
    for s in rows:
        tiers[s["tier"]] += 1
    out = {"leads": len(rows), "tier_a": tiers["A"], "tier_b": tiers["B"], "tier_c": tiers["C"], "dq": tiers["DQ"]}
    if tiers["D"]:
        out["tier_d"] = tiers["D"]
    return out


def top_hooks(conn: sqlite3.Connection, run_id: str, k: int = 3) -> list[str]:
    counts: dict[str, int] = {}
    for s in db.scores_for_run(conn, run_id):
        hooks = json.loads(s["need_hooks_json"])
        if hooks:
            counts[hooks[0]] = counts.get(hooks[0], 0) + 1
    return [f"{n}× {h}" for h, n in sorted(counts.items(), key=lambda x: -x[1])[:k]]


# used by export test to keep openpyxl import honest
def _today() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _stale_flag(verified_iso: str, staleness_days: int) -> str:
    """'yes' when the newest evidence is older than validation.staleness_days (docs/03 §staleness).
    Never-verified and unparseable timestamps say so — '-' must only ever mean 'fresh'."""
    if not verified_iso:
        return "never verified"
    from datetime import datetime, timedelta
    try:
        seen = datetime.fromisoformat(verified_iso.replace("Z", "+00:00"))
        if seen.tzinfo is None:
            seen = seen.replace(tzinfo=UTC)
    except ValueError:
        return "unknown (bad timestamp)"
    return "yes" if datetime.now(tz=UTC) - seen > timedelta(days=staleness_days) else ""


_DAY_ORDER = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")


def _format_hours(hours_json: str | None) -> str:
    """'Mon 9AM-6PM | ... | Sun closed' from the listing's hours dict; '-' when unknown."""
    if not hours_json:
        return "-"
    try:
        hours = json.loads(hours_json)
    except (TypeError, ValueError):
        return "-"
    if not isinstance(hours, dict) or not hours:
        return "-"
    parts = []
    for day in _DAY_ORDER:
        vals = hours.get(day)
        if not vals:
            continue
        txt = ", ".join(vals) if isinstance(vals, list) else str(vals)
        txt = txt.replace(" ", " ").replace("–", "-").replace(" ", " ")
        parts.append(f"{day[:3]} {txt}")
    return " | ".join(parts) if parts else "-"
