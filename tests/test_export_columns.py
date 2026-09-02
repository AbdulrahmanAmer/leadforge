"""v0.3 D — export truth: the new default-profile columns are present, non-blank (zero-blank rule),
Next Action is phone-first, and a freemail address never outranks an own-domain one in the Email
column (docs/09 §D acceptance)."""
import csv

from leadforge import db
from leadforge.config import load_config
from leadforge.export import ACCOUNT_COLUMNS, DEFAULT_EXTRA_COLUMNS, _site_status, export_run
from leadforge.models import ICP, Business, Contact, Person, Score, ScoreFactor


def _icp():
    return ICP.model_validate({"campaign": "t", "offer": {"what": "web design"},
                               "target": {"categories": ["auto repair shop"],
                                          "geography": {"areas": ["Houston, TX"], "country": "US"}}})


def _bootstrap(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    return cfg, conn, rid


def _score(conn, rid, bid):
    db.save_score(conn, Score(business_id=bid, run_id=rid, total=70, tier="B",
                              factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                   points=1, why="w")]))


def _csv_rows(arts):
    path = next(a for a in arts if a.endswith(".csv"))
    with open(path, encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def test_new_columns_present_and_never_blank(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Bare Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    for col in DEFAULT_EXTRA_COLUMNS:
        assert col in row, f"missing column {col}"
        assert row[col] not in ("", None), f"{col} is blank on a bare-minimum row"
    # a bare business with no score-side meta factors still gets an honest placeholder, not a crash
    assert row["Contactability"] == "not computed"
    assert row["Status"] == "not computed"
    assert row["Entity Type"] == "unchecked"
    assert row["Chain"] == "-"


def test_account_fit_export_unaffected_by_the_new_columns(tmp_path, monkeypatch):
    """account_fit rows keep their own column set — the v0.3 D columns are additive to the default
    profile only, per the unit brief ('Keep account_fit untouched except sharing helpers')."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    icp = ICP.model_validate({"campaign": "t", "offer": {"what": "x"},
                              "target": {"categories": ["manufacturer"],
                                         "geography": {"areas": ["Dubai"], "country": "AE"}},
                              "scoring": {"profile": "account_fit"}})
    db.upsert_business(conn, Business(id="b1", name="Acme", source="gosom", dedupe_key="dk-b1"))
    db.save_score(conn, Score(business_id="b1", run_id=rid, total=70, tier="B",
                              factors=[ScoreFactor(factor="status", group="meta", weight=0, score=0,
                                                   points=0, why="NEW")]))
    arts = export_run(conn, icp, rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    for col in ("Fit", "Next Action", "Entity Type", "Lawful Basis (Email)"):
        assert col not in row
    assert "Status" in row and row["Status"] == "NEW"
    assert set(ACCOUNT_COLUMNS) == set(row.keys())


def test_next_action_is_phone_first(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="named", name="Named Garage", source="gosom", dedupe_key="dk-named",
                                      phone_e164="+441483123456"))
    db.add_person(conn, Person(business_id="named", name="Jo Owner", title="Owner", is_dm=1,
                               dm_confidence=0.9, labeled_by="agent"))
    _score(conn, rid, "named")

    db.upsert_business(conn, Business(id="switch", name="Switch Garage", source="gosom",
                                      dedupe_key="dk-switch", phone_e164="+441483123457"))
    _score(conn, rid, "switch")

    db.upsert_business(conn, Business(id="mail", name="Mail Garage", source="gosom", dedupe_key="dk-mail",
                                      domain="mailgarage.example"))
    db.add_contact(conn, Contact(business_id="mail", kind="email", value="info@mailgarage.example",
                                 tier="valid", affinity="own_domain"))
    _score(conn, rid, "mail")

    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    rows = {r["Business"]: r for r in _csv_rows(arts)}
    # a validated phone always wins Next Action even when a DM is present (phone-first, owner decision 2)
    assert rows["Named Garage"]["Next Action"] == "CALL - named contact"
    assert rows["Switch Garage"]["Next Action"] == "CALL - ask for the owner"
    assert rows["Mail Garage"]["Next Action"] == "EMAIL - eligible"


def test_next_action_shows_outreach_state_when_an_outreach_targets_row_exists(tmp_path, monkeypatch):
    """v0.3 polish finding 5: once a lead is enrolled in a campaign, Next Action shows the live
    outreach lifecycle state ('OUTREACH - <state>') instead of the phone-first default — even for a
    business that would otherwise read 'CALL - named contact'."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="enrolled", name="Enrolled Garage", source="gosom",
                                      dedupe_key="dk-enrolled", phone_e164="+441483123456"))
    db.add_person(conn, Person(business_id="enrolled", name="Jo Owner", title="Owner", is_dm=1,
                               dm_confidence=0.9, labeled_by="agent"))
    _score(conn, rid, "enrolled")
    conn.execute(
        "INSERT INTO outreach_targets(business_id,campaign,state,created_at,updated_at) "
        "VALUES(?,?,?,?,?)", ("enrolled", "t", "approved", "2026-01-01T00:00:00Z", "2026-01-01T00:00:00Z"),
    )
    conn.commit()
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Next Action"] == "OUTREACH - approved"


def test_email_confidence_agrees_with_lawful_basis_on_legacy_rows(tmp_path, monkeypatch):
    """Real-data proof caught this: a pre-v0.3 contact (no `affinity` stored) showed Lawful Basis
    'b2b_legitimate_interest' but Email Confidence 'none' for the SAME address — two readings of one
    fact. Both must now agree that a legacy own-domain role address is exactly that."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="legacy", name="Legacy Garage", source="gosom",
                                      dedupe_key="dk-legacy", domain="legacygarage.example"))
    db.add_contact(conn, Contact(business_id="legacy", kind="email", value="info@legacygarage.example",
                                 tier="role"))  # affinity left unset, as every pre-v0.3 row is
    _score(conn, rid, "legacy")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Email Confidence"] == "own domain, role mailbox"
    assert row["Lawful Basis (Email)"] == "b2b_legitimate_interest"


def test_freemail_never_outranks_own_domain_in_the_email_column(tmp_path, monkeypatch):
    """The v0.2 sheet exported a font designer's gmail above a real info@ three times — this is the
    regression test for the fix: a freemail 'valid' address must never beat an own-domain 'role' one."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="dual", name="Dual Garage", source="gosom", dedupe_key="dk-dual",
                                      domain="dualgarage.example"))
    db.add_contact(conn, Contact(business_id="dual", kind="email", value="jane.font@gmail.com",
                                 tier="valid", affinity="freemail_linked"))
    db.add_contact(conn, Contact(business_id="dual", kind="email", value="info@dualgarage.example",
                                 tier="role", affinity="own_domain"))
    _score(conn, rid, "dual")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Email"] == "info@dualgarage.example"
    assert row["Email Tier"] == "role"
    assert row["Email Confidence"] == "own domain, role mailbox"


def test_freemail_never_outranks_own_domain_when_affinity_is_unset_on_both(tmp_path, monkeypatch):
    """The real-data shape (fresh-context review blocker): 233/233 contact rows on the live campaign
    DB store affinity '' — BOTH the freemail and the own-domain contact, not just one. Ranking on the
    raw rows falls through to tier order alone and the freemail 'valid' address beats the own-domain
    'role' one; affinity must be backfilled BEFORE ranking, not only on the already-chosen winner."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="legacy2", name="Legacy2 Garage", source="gosom",
                                      dedupe_key="dk-legacy2", domain="legacy2garage.example"))
    db.add_contact(conn, Contact(business_id="legacy2", kind="email", value="dave.font@gmail.com",
                                 tier="valid"))  # affinity unset, as every pre-v0.3 row is
    db.add_contact(conn, Contact(business_id="legacy2", kind="email", value="info@legacy2garage.example",
                                 tier="role"))  # affinity unset here too
    _score(conn, rid, "legacy2")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Email"] == "info@legacy2garage.example"
    assert row["Email Confidence"] == "own domain, role mailbox"
    assert row["Lawful Basis (Email)"] == "b2b_legitimate_interest"


def test_email_confidence_admits_when_freemail_linkage_was_never_checked(tmp_path, monkeypatch):
    """A legacy row's freemail affinity is a coarse domain-only guess (any freemail box == 'linked'),
    not the real name-token linkage check — the column must say so, not claim a check that never ran."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="onlyfree", name="Only Freemail Garage", source="gosom",
                                      dedupe_key="dk-onlyfree", domain="onlyfreegarage.example"))
    db.add_contact(conn, Contact(business_id="onlyfree", kind="email", value="stranger@gmail.com",
                                 tier="valid"))  # affinity unset -> backfilled, not a real linkage check
    _score(conn, rid, "onlyfree")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Email"] == "stranger@gmail.com"
    assert row["Email Confidence"] == "personal freemail (linkage not checked — pre-v0.3 row)"


def test_why_this_score_never_shows_the_contactability_meta_factor(tmp_path, monkeypatch):
    """'Why This Score' explains Score/Fit — contactability is a separate meta axis (docs/09 §D split)
    that can carry up to 98 points, dwarfing every real fit factor (max 25); it must never be able to
    push into the top-3 'Why This Score' drivers just because it happens to score the most points."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="whyrow", name="Why Garage", source="gosom", dedupe_key="dk-whyrow"))
    db.save_score(conn, Score(business_id="whyrow", run_id=rid, total=25, tier="C", factors=[
        ScoreFactor(factor="industry_match", group="fit", weight=25, score=1.0, points=25,
                   why="category matches the ICP"),
        ScoreFactor(factor="contactability", group="meta", weight=100, score=0.98, points=98,
                   why="DM identified; own-domain valid email; validated phone"),
        ScoreFactor(factor="status", group="meta", weight=0, score=0.0, points=0, why="READY"),
    ]))
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Why This Score"] == "category matches the ICP"
    assert "DM identified" not in row["Why This Score"]


def test_site_status_robots_disallowed_is_not_dead():
    """Real data: 'dead (robots-disallowed)' on 5 live-campaign rows was a lie — the crawler was
    refused, the site is not down. site_dead must stay False so email_eligibility is never blocked
    for a reason that was never actually observed."""
    enrich = {"crawled_at": "2026-01-01T00:00:00Z", "error": "robots-disallowed", "signals": {}}
    status, dead = _site_status(enrich)
    assert status == "not crawlable (robots)"
    assert dead is False


def test_site_status_still_reports_dead_for_a_real_failure():
    enrich = {"crawled_at": "2026-01-01T00:00:00Z", "error": "home unreachable (status:404)",
             "signals": {"http_status": 404}}
    status, dead = _site_status(enrich)
    assert status == "dead (404)"
    assert dead is True


# --- v0.3 polish finding 3/1: the full decision table, in the order it is checked -------------
def test_site_status_live_when_crawled_with_pages():
    enrich = {"crawled_at": "2026-01-01T00:00:00Z", "pages": 3, "signals": {}}
    assert _site_status(enrich) == ("live", False)


def test_site_status_redirect_when_crawled_with_pages_and_offsite():
    enrich = {"crawled_at": "2026-01-01T00:00:00Z", "pages": 1,
             "signals": {"offsite_redirect": True, "final_host": "newsite.example"}}
    assert _site_status(enrich) == ("redirects to newsite.example", False)


def test_site_status_phantom_crawl_is_not_live():
    """v0.3 polish finding 1: crawled_at stamped but pages=0 (the live campaign's 115 'phantom'
    crawls) must NOT read 'live' — checking crawled_at alone was the bug."""
    enrich = {"crawled_at": "2026-01-01T00:00:00Z", "pages": 0, "signals": {}}
    status, dead = _site_status(enrich)
    assert status != "live"
    assert status == "not crawled"
    assert dead is False


def test_site_status_unreachable_when_attempted_but_never_crawled_and_no_error():
    """attempted_at without crawled_at and without an error: the crawler tried and recorded
    nothing usable back — distinct from both a real failure (has an error) and 'not crawled'
    (was never even attempted)."""
    enrich = {"attempted_at": "2026-01-01T00:00:00Z", "signals": {}}
    assert _site_status(enrich) == ("unreachable", True)


def test_site_status_not_crawled_when_nothing_recorded_at_all():
    assert _site_status({}) == ("not crawled", False)


# --- v0.3 polish finding 5: Summary funnel + Next Action breakdown, read back from the workbook -
def test_summary_sheet_funnel_and_next_action_breakdown(tmp_path, monkeypatch):
    """The Summary sheet's funnel counts and Next Action breakdown are read back from the actual
    XLSX (openpyxl), not asserted against internals — proof the numbers a human opens are correct,
    not just the row dicts they were built from."""
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    # 1) call-ready: named DM + validated phone -> Next Action "CALL - named contact"
    db.upsert_business(conn, Business(id="named", name="Named Garage", source="gosom",
                                      dedupe_key="dk-named", phone_e164="+441483123456"))
    db.add_person(conn, Person(business_id="named", name="Jo Owner", title="Owner", is_dm=1,
                               dm_confidence=0.9, labeled_by="agent"))
    _score(conn, rid, "named")
    # 2) email-eligible, own-domain, no phone -> Next Action "EMAIL - eligible"
    db.upsert_business(conn, Business(id="mail", name="Mail Garage", source="gosom", dedupe_key="dk-mail",
                                      website="https://mailgarage.example", domain="mailgarage.example"))
    db.add_contact(conn, Contact(business_id="mail", kind="email", value="info@mailgarage.example",
                                 tier="valid", affinity="own_domain"))
    _score(conn, rid, "mail")
    # 3) nothing at all -> Next Action "RESEARCH - no reachable channel"
    db.upsert_business(conn, Business(id="bare", name="Bare Garage", source="gosom", dedupe_key="dk-bare"))
    _score(conn, rid, "bare")

    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    wb = load_workbook(next(a for a in arts if a.endswith(".xlsx")))
    ws = wb["Summary"]
    lines = {row[0].value: row[1].value for row in ws.iter_rows(min_col=1, max_col=2) if row[0].value}

    assert lines["Total leads"] == 3
    assert lines["  With website"] == "1 (33%)"          # only "mail"
    assert lines["  With any email"] == "1 (33%)"
    assert lines["  With own-domain email"] == "1 (33%)"
    assert lines["  Eligible to email"] == "1 (33%)"
    assert lines["  Call-ready (validated phone)"] == "1 (33%)"  # only "named"

    # Next Action breakdown rows: col A = "  <n>×", col B = the Next Action label — every row here
    # has count 1 (one business per bucket), so read them as (label -> count) pairs, not a dict
    # keyed by the count label (three rows would collide on the same "  1×" key).
    na_counts = {row[1].value: int(str(row[0].value).strip().rstrip("×"))
                for row in ws.iter_rows(min_col=1, max_col=2)
                if isinstance(row[0].value, str) and row[0].value.strip().rstrip("×").isdigit()}
    assert na_counts["CALL - named contact"] == 1
    assert na_counts["EMAIL - eligible"] == 1
    assert na_counts["RESEARCH - no reachable channel"] == 1
