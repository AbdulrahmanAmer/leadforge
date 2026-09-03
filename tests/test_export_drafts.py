"""v0.4 "autopilot" C — the exported sheet carries the drafted emails (docs/08 ADR-015). Nothing here
is ever sent: sending stays behind `leadforge outreach approve` / `--live`, unchanged by this unit.

`messages.author` does not exist in this worktree's schema (v2) — B adds it in schema v3, in
parallel. export.py must read it defensively (`m["author"] if "author" in m.keys() else "agent"`), so
these tests insert `messages` rows WITHOUT an `author` column, on purpose, to prove that path works.
"""
import csv
import json

from leadforge import db
from leadforge.config import load_config
from leadforge.export import ACCOUNT_COLUMNS, DEFAULT_EXTRA_COLUMNS, DRAFT_COLUMNS, export_run
from leadforge.models import ICP, Business, Contact, Person, Score, ScoreFactor
from leadforge.util import now_iso


def _icp(campaign="t"):
    return ICP.model_validate({"campaign": campaign, "offer": {"what": "web design"},
                               "target": {"categories": ["auto repair shop"],
                                          "geography": {"areas": ["Houston, TX"], "country": "US"}}})


def _bootstrap(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    return cfg, conn, rid


def _score(conn, rid, bid):
    db.save_score(conn, Score(business_id=bid, run_id=rid, total=70, tier="B",
                              factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                   points=1, why="w")]))


def _csv_rows(arts):
    path = next(a for a in arts if a.endswith(".csv"))
    with open(path, encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def _target(conn, business_id, campaign, contact_id=None, state="drafted"):
    now = now_iso()
    conn.execute(
        "INSERT INTO outreach_targets(business_id,contact_id,campaign,state,created_at,updated_at) "
        "VALUES(?,?,?,?,?,?)", (business_id, contact_id, campaign, state, now, now),
    )
    conn.commit()
    row = conn.execute("SELECT id FROM outreach_targets WHERE business_id=? AND campaign=?",
                       (business_id, campaign)).fetchone()
    return row["id"]


def _message(conn, target_id, *, subject="Quick note", body="Hi there.", grade="A",
            used_fact="site_stale", state="drafted", step=1):
    """Insert a messages row WITHOUT an `author` column — this worktree's schema (v2) doesn't have
    one yet; export.py must fall back to 'agent' when reading it."""
    now = now_iso()
    conn.execute(
        "INSERT INTO messages(target_id,step,purpose,subject,body_text,draft_hash,state,gate_json,"
        "grade,used_fact,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
        (target_id, step, "cold_open", subject, body, "hash", state, "{}", grade, used_fact, now, now),
    )
    conn.commit()


def test_new_columns_present_and_never_blank_include_draft_columns(tmp_path, monkeypatch):
    """DRAFT_COLUMNS ride inside DEFAULT_EXTRA_COLUMNS, so the existing zero-blank contract already
    covers them — a bare business with no message at all still shows 'no draft', never blank."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Bare Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    for col in DRAFT_COLUMNS:
        assert col in DEFAULT_EXTRA_COLUMNS
        assert col in row, f"missing column {col}"
    assert row["Draft Subject"] == "no draft"
    assert row["Draft Body"] == "no draft"
    assert row["Draft Grade"] == "no draft"
    assert row["Draft Author"] == "no draft"


def test_csv_shows_the_four_draft_columns_from_a_seeded_message(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Drafted Garage", source="gosom", dedupe_key="dk-b1",
                                      domain="drafted.example"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t")
    _message(conn, tid, subject="Noticed your site", body="Full body text here.", grade="B",
             used_fact="site_stale")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Draft Subject"] == "Noticed your site"
    assert row["Draft Body"] == "Full body text here."
    assert row["Draft Grade"] == "B"
    assert row["Draft Author"] == "agent"  # no `author` column in this worktree's schema -> fallback


def test_rejected_only_message_shows_no_draft(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Rejected Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t")
    _message(conn, tid, subject="Rejected subject", body="Rejected body", grade="C",
             used_fact="", state="rejected")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Draft Subject"] == "no draft"
    assert row["Draft Body"] == "no draft"
    assert row["Draft Grade"] == "no draft"
    assert row["Draft Author"] == "no draft"


def test_newest_non_rejected_message_wins_over_an_older_one(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Two Draft Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t")
    _message(conn, tid, subject="Old subject", body="Old body", grade="A", used_fact="hiring", step=1)
    _message(conn, tid, subject="New subject", body="New body", grade="B", used_fact="rating", step=2)
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Draft Subject"] == "New subject"
    assert row["Draft Grade"] == "B"


def test_message_in_a_different_campaign_is_not_picked_up(tmp_path, monkeypatch):
    """outreach_targets is unique on (business_id, campaign) — a business enrolled under a DIFFERENT
    campaign's messages must never leak into this campaign's export."""
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Other Campaign Garage", source="gosom",
                                      dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "other-campaign")
    _message(conn, tid, subject="Wrong campaign subject")
    arts = export_run(conn, _icp(campaign="t"), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    assert row["Draft Subject"] == "no draft"


def test_account_fit_unaffected_by_draft_columns(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    icp = ICP.model_validate({"campaign": "t", "offer": {"what": "x"},
                              "target": {"categories": ["manufacturer"],
                                         "geography": {"areas": ["Dubai"], "country": "AE"}},
                              "scoring": {"profile": "account_fit"}})
    db.upsert_business(conn, Business(id="b1", name="Acme", source="gosom", dedupe_key="dk-b1"))
    db.save_score(conn, Score(business_id="b1", run_id=rid, total=70, tier="B",
                              factors=[ScoreFactor(factor="status", group="meta", weight=0, score=0,
                                                   points=0, why="NEW")]))
    arts = export_run(conn, icp, rid, cfg.exports_dir, ["csv"], cfg=cfg)
    row = _csv_rows(arts)[0]
    for col in DRAFT_COLUMNS:
        assert col not in row
    assert set(ACCOUNT_COLUMNS) == set(row.keys())


def test_xlsx_has_drafts_sheet_with_one_row_and_wrapped_body(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Sheet Garage", source="gosom", dedupe_key="dk-b1",
                                      domain="sheetgarage.example"))
    db.add_person(conn, Person(business_id="b1", name="Jo Owner", title="Owner", is_dm=1,
                               dm_confidence=0.9, labeled_by="agent"))
    contact_id = db.add_contact(conn, Contact(business_id="b1", kind="email",
                                              value="jo@sheetgarage.example", tier="valid",
                                              affinity="own_domain"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t", contact_id=contact_id)
    _message(conn, tid, subject="Sheet subject", body="Sheet body text.", grade="A",
             used_fact="booking")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    xlsx_path = next(a for a in arts if a.endswith(".xlsx"))
    wb = load_workbook(xlsx_path)
    assert wb.sheetnames.index("Drafts") == wb.sheetnames.index("Leads") + 1  # right after Leads
    ws = wb["Drafts"]
    header = [c.value for c in ws[1]]
    assert header == ["Business", "DM Name", "To", "Subject", "Body", "Grade", "Used Fact", "Author",
                      "State", "Created"]
    data_row = [c.value for c in ws[2]]
    assert data_row[0] == "Sheet Garage"
    assert data_row[1] == "Jo Owner"
    assert data_row[2] == "jo@sheetgarage.example"  # the contact_id's own value, not a re-ranked one
    assert data_row[3] == "Sheet subject"
    assert data_row[4] == "Sheet body text."
    assert data_row[5] == "A"
    assert data_row[6] == "booking"
    assert data_row[7] == "agent"
    assert data_row[8] == "drafted"
    body_cell = ws.cell(row=2, column=5)
    assert body_cell.alignment.wrap_text is True
    assert ws.column_dimensions["E"].width >= 60  # wide body column


def test_xlsx_drafts_sheet_falls_back_to_best_email_when_no_contact_id(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="No Contact Garage", source="gosom",
                                      dedupe_key="dk-b1", domain="nocontact.example"))
    db.add_contact(conn, Contact(business_id="b1", kind="email", value="info@nocontact.example",
                                 tier="role", affinity="own_domain"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t")  # no contact_id
    _message(conn, tid, subject="Fallback subject")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    wb = load_workbook(next(a for a in arts if a.endswith(".xlsx")))
    ws = wb["Drafts"]
    assert [c.value for c in ws[2]][2] == "info@nocontact.example"


def test_xlsx_drafts_sheet_absent_when_no_messages_at_all(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="No Draft Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    wb = load_workbook(next(a for a in arts if a.endswith(".xlsx")))
    ws = wb["Drafts"]
    assert ws.max_row == 1  # header only, no data rows


def test_leads_sheet_draft_body_cell_is_wrapped(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Wrap Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    tid = _target(conn, "b1", "t")
    _message(conn, tid, subject="Wrap subject", body="Wrap body text.")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    wb = load_workbook(next(a for a in arts if a.endswith(".xlsx")))
    ws = wb["Leads"]
    header = [c.value for c in ws[1]]
    body_col = header.index("Draft Body") + 1
    assert ws.cell(row=2, column=body_col).value == "Wrap body text."
    assert ws.cell(row=2, column=body_col).alignment.wrap_text is True


def test_about_sheet_explains_draft_columns(tmp_path, monkeypatch):
    from openpyxl import load_workbook

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="About Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["xlsx"], cfg=cfg)
    wb = load_workbook(next(a for a in arts if a.endswith(".xlsx")))
    ws = wb["About"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "never sent" in text
    assert "leadforge outreach approve" in text
    assert "--live" in text


def test_report_json_and_digest_count_drafted_and_by_author(tmp_path, monkeypatch):
    from leadforge.export import summarize_for_digest

    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="drafted1", name="Drafted One", source="gosom",
                                      dedupe_key="dk-d1"))
    _score(conn, rid, "drafted1")
    t1 = _target(conn, "drafted1", "t")
    _message(conn, t1, subject="s1")

    db.upsert_business(conn, Business(id="drafted2", name="Drafted Two", source="gosom",
                                      dedupe_key="dk-d2"))
    _score(conn, rid, "drafted2")
    t2 = _target(conn, "drafted2", "t")
    _message(conn, t2, subject="s2")

    db.upsert_business(conn, Business(id="none1", name="No Draft", source="gosom", dedupe_key="dk-n1"))
    _score(conn, rid, "none1")

    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    report = json.loads(open(next(a for a in arts if a.endswith("report.json")),
                            encoding="utf-8").read())
    assert report["drafted"] == 2
    assert report["drafts_by_author"] == {"agent": 2}

    digest = summarize_for_digest(conn, rid)
    assert digest["drafted"] == 2
    assert digest["drafts_by_author"] == {"agent": 2}


def test_report_json_drafted_is_zero_with_no_messages(tmp_path, monkeypatch):
    cfg, conn, rid = _bootstrap(tmp_path, monkeypatch)
    db.upsert_business(conn, Business(id="b1", name="Empty Garage", source="gosom", dedupe_key="dk-b1"))
    _score(conn, rid, "b1")
    arts = export_run(conn, _icp(), rid, cfg.exports_dir, ["csv"], cfg=cfg)
    report = json.loads(open(next(a for a in arts if a.endswith("report.json")),
                            encoding="utf-8").read())
    assert report["drafted"] == 0
    assert report["drafts_by_author"] == {}
