"""v0.1.4: summary/report stats count only REAL data — placeholder cells are for humans on the
phone, not for coverage metrics (live 709-run reported with_dm=709 when 330 were placeholders) —
and registry 'SURNAME, Given Names' officers read naturally on the sheet."""
import csv
import json

from leadforge import db
from leadforge.util import natural_name


def _minimal_icp():
    from leadforge.models import ICP
    return ICP.model_validate({"campaign": "t", "offer": {"what": "x"},
                               "target": {"categories": ["garage"],
                                          "geography": {"areas": ["Guildford"], "country": "GB"}}})


def _seeded_run(tmp_path, monkeypatch):
    """Two scored businesses: one with a real registry DM + valid email, one with nothing."""
    from leadforge.config import load_config
    from leadforge.models import Business, Contact, Person, Score, ScoreFactor
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    db.upsert_business(conn, Business(id="real", run_id=rid, name="Real Garage", source="gosom",
                                      website="https://real.example", phone_e164="+441483123456",
                                      dedupe_key="dk-real"))
    db.upsert_business(conn, Business(id="bare", run_id=rid, name="Bare Garage", source="gosom",
                                      dedupe_key="dk-bare"))
    db.add_person(conn, Person(business_id="real", name="Murphy, Sean Vincent", title="Director",
                               labeled_by="registry", is_dm=1, dm_confidence=0.9))
    db.add_contact(conn, Contact(business_id="real", kind="email", value="sean@real.example",
                                 label="direct", tier="valid"))
    for bid in ("real", "bare"):
        db.save_score(conn, Score(business_id=bid, run_id=rid, total=60, tier="B",
                                  factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                       points=1, why="w")]))
    return cfg, conn, rid


def test_report_counts_only_real_dm_and_email(tmp_path, monkeypatch):
    from leadforge.export import export_run
    cfg, conn, rid = _seeded_run(tmp_path, monkeypatch)
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["csv"])
    report_path = next(a for a in arts if a.endswith("report.json"))
    report = json.loads(open(report_path, encoding="utf-8").read())
    assert report["total"] == 2
    assert report["with_dm"] == 1      # the "not identified" placeholder row must not count
    assert report["with_email"] == 1   # nor "no website to crawl"


def test_sheet_stays_fully_resolved_and_private_keys_stay_private(tmp_path, monkeypatch):
    from leadforge.export import export_run
    cfg, conn, rid = _seeded_run(tmp_path, monkeypatch)
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["csv"])
    csv_path = next(a for a in arts if a.endswith(".csv"))
    with open(csv_path, encoding="utf-8-sig") as fh:
        rdr = csv.DictReader(fh)
        rows = list(rdr)
        assert not [k for k in rdr.fieldnames if k.startswith("_")]  # honesty flags never exported
    named = next(r for r in rows if r["Business"] == "Real Garage")
    assert named["DM Name"] == "Sean Vincent Murphy"  # natural order on the sheet, even from old DBs
    assert named["Call Readiness"] == "READY - named contact"
    bare = next(r for r in rows if r["Business"] == "Bare Garage")
    assert bare["DM Name"].startswith("not identified")  # zero-blank-cell rule unchanged
    assert bare["Email"] == "no website to crawl"


def test_export_neutralizes_formula_injection_and_control_chars(tmp_path, monkeypatch):
    """A scraped business name must never execute in the operator's Excel, and control chars
    must not crash openpyxl at the very end of a long run."""
    from openpyxl import load_workbook

    from leadforge.config import load_config
    from leadforge.export import export_run
    from leadforge.models import Business, Score, ScoreFactor
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    db.upsert_business(conn, Business(id="evil", run_id=rid, source="gosom", dedupe_key="dk-evil",
                                      name='=HYPERLINK("http://evil.example","click")'))
    db.upsert_business(conn, Business(id="ctrl", run_id=rid, source="gosom", dedupe_key="dk-ctrl",
                                      name="Null\x00 Garage\x07"))
    for bid in ("evil", "ctrl"):
        db.save_score(conn, Score(business_id=bid, run_id=rid, total=60, tier="B",
                                  factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                       points=1, why="w")]))
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["xlsx", "csv"])  # must not raise
    with open(next(a for a in arts if a.endswith(".csv")), encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    evil = next(r for r in rows if "HYPERLINK" in r["Business"])
    assert evil["Business"].startswith("'=")  # neutralized, not live
    ctrl = next(r for r in rows if "Garage" in r["Business"])
    assert ctrl["Business"] == "Null Garage"
    ws = load_workbook(next(a for a in arts if a.endswith(".xlsx")))["Leads"]
    biz_col = [c.value for c in ws[1]].index("Business") + 1
    cells = [ws.cell(row=i, column=biz_col) for i in range(2, 4)]
    assert all(c.data_type != "f" for c in cells)  # no formula cells anywhere
    web_col = [c.value for c in ws[1]].index("Website") + 1
    for i in (2, 3):  # both rows have no website: placeholder text must NOT be a hyperlink
        assert ws.cell(row=i, column=web_col).hyperlink is None


def test_tier_d_is_counted_not_dropped(tmp_path, monkeypatch):
    """account_fit grades A-D; Summary/report used to enumerate only A/B/C/DQ, silently losing D."""
    import json as _json

    from leadforge.config import load_config
    from leadforge.export import export_run
    from leadforge.models import Business, Score, ScoreFactor
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    db.upsert_business(conn, Business(id="d1", run_id=rid, name="Grade D Garage", source="gosom",
                                      dedupe_key="dk-d1"))
    db.save_score(conn, Score(business_id="d1", run_id=rid, total=30, tier="D",
                              factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                   points=1, why="w")]))
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["csv"])
    report = _json.loads(open(next(a for a in arts if a.endswith("report.json")), encoding="utf-8").read())
    assert report["tiers"].get("D") == 1
    assert sum(report["tiers"].values()) == report["total"]


def test_stale_flag_and_call_readiness_tell_the_truth(tmp_path, monkeypatch):
    """v0.1.4: Stale? '' meant fresh AND never-verified AND bad-timestamp; a raw unparsed Maps
    phone string counted as call-ready."""
    from leadforge.config import load_config
    from leadforge.export import export_run
    from leadforge.models import Business, Score, ScoreFactor
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    db.upsert_business(conn, Business(id="raw", run_id=rid, name="Raw Phone Garage", source="gosom",
                                      dedupe_key="dk-raw", phone_raw="Call us: O800-CARS"))
    db.save_score(conn, Score(business_id="raw", run_id=rid, total=60, tier="B",
                              factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                   points=1, why="w")]))
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["csv"])
    with open(next(a for a in arts if a.endswith(".csv")), encoding="utf-8-sig") as fh:
        row = next(iter(csv.DictReader(fh)))
    assert row["Stale?"] == "never verified"          # no evidence ever collected
    assert row["Call Readiness"] == "UNVERIFIED PHONE - confirm number"
    assert row["Phone"] == "Call us: O800-CARS"       # still displayed — honesty, not deletion


def test_natural_name_rules():
    assert natural_name("Murphy, Sean Vincent") == "Sean Vincent Murphy"
    assert natural_name("MURPHY, Sean") == "Sean MURPHY"              # casing is the caller's job
    assert natural_name("Jane Smith") == "Jane Smith"                 # already natural
    assert natural_name("Acme Widgets, Inc") == "Acme Widgets, Inc"   # corporate comma stays
    assert natural_name("SMITH, John, Jr") == "SMITH, John, Jr"       # two commas: don't guess
    assert natural_name(" SMITH ,  Jane ") == "Jane SMITH"
    assert natural_name("") == ""


def test_inferred_email_never_masquerades_as_a_found_one(tmp_path, monkeypatch):
    """v0.2.0 honesty contract: a guess lives in its own column, is labeled as likely, and is
    NEVER counted in the published-email coverage figure a buyer reads."""
    import json as _json

    from leadforge.config import load_config
    from leadforge.export import export_run
    from leadforge.models import Business, Contact, Person, Score, ScoreFactor
    monkeypatch.chdir(tmp_path)
    cfg = load_config(tmp_path)
    conn = db.connect(cfg.db_path)
    rid = db.create_run(conn, "icp.yaml", "h")
    db.upsert_business(conn, Business(id="g", run_id=rid, name="Guessed Garage", source="gosom",
                                      dedupe_key="dk-g", website="https://g.example", domain="g.example"))
    db.add_person(conn, Person(business_id="g", name="Jane Smith", title="Director",
                               labeled_by="registry", is_dm=1, dm_confidence=0.9))
    db.add_contact(conn, Contact(business_id="g", kind="email", value="jane.smith@g.example",
                                 label="inferred", tier="inferred",
                                 meta={"pattern": "first.last", "confidence": 0.45,
                                       "basis": "pattern first.last from bob.jones@g.example"}))
    db.save_score(conn, Score(business_id="g", run_id=rid, total=60, tier="B",
                              factors=[ScoreFactor(factor="x", group="fit", weight=1, score=1,
                                                   points=1, why="w")]))
    arts = export_run(conn, _minimal_icp(), rid, cfg.exports_dir, ["csv"])
    with open(next(a for a in arts if a.endswith(".csv")), encoding="utf-8-sig") as fh:
        row = next(iter(csv.DictReader(fh)))
    assert "jane.smith@g.example" in row["Email (Inferred)"]
    assert "likely" in row["Email (Inferred)"]           # the cell carries its own caveat
    assert "jane.smith@g.example" not in row["Email"]    # never leaks into the mail-merge column
    assert row["Email"] == "site not crawled"            # still says why there is no real address
    report = _json.loads(open(next(a for a in arts if a.endswith("report.json")), encoding="utf-8").read())
    assert report["with_email"] == 0                      # a guess is not coverage
    assert report["with_inferred_email"] == 1             # counted, separately and honestly
